from openpyxl import load_workbook

from vasp_auto.excel_writer import write_results_to_excel


def test_write_results_to_excel_styles_converged(tmp_path):
    excel_path = tmp_path / "summary.xlsx"
    rows = [
        {"project": "p", "case": "good", "converged": True, "energy_eV": -12.5},
        {"project": "p", "case": "bad", "converged": False, "energy_eV": None},
    ]

    write_results_to_excel(str(excel_path), rows)

    worksheet = load_workbook(excel_path)["results"]
    header = [cell.value for cell in worksheet[1]]
    converged_col = header.index("converged") + 1
    assert worksheet.cell(row=2, column=converged_col).fill.start_color.rgb.endswith("C6EFCE")
    assert worksheet.cell(row=3, column=converged_col).fill.start_color.rgb.endswith("FFC7CE")


def test_write_results_to_excel_empty(tmp_path):
    excel_path = tmp_path / "empty.xlsx"
    write_results_to_excel(str(excel_path), [])
    assert excel_path.exists()
