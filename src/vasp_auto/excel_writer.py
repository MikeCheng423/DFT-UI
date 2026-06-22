from pathlib import Path
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


# Preferred ordering for the columns the row builders actually emit; anything
# unknown is appended after these, and job_dir always goes last.
COLUMN_ORDER = [
    "project",
    "case",
    "engine",
    "calculation_type",
    "step",
    "status",
    "converged",
    "energy_eV",
    "fermi_eV",
    "band_gap_eV",
    "max_force_eV_A",
    "pressure_kB",
    "magmom_total",
    "magmoms",
    "neb_barrier_eV",
    "neb_forward_barrier_eV",
    "neb_backward_barrier_eV",
    "neb_image_energies_eV",
    "ionic_steps",
    "selected_encut",
    "selected_sigma",
    "selected_nelm",
    "selected_kpoints",
    "errors",
    "auto_retries",
    "auto_fixes",
    "return_code",
    "job_id",
    "convergence_report",
    "convergence_csv",
    "job_dir",
]

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MAX_COLUMN_WIDTH = 60


def _style_worksheet(worksheet, df: pd.DataFrame):
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = HEADER_FILL

    if "converged" in df.columns:
        converged_col = df.columns.get_loc("converged") + 1
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=converged_col)
            if cell.value is True or str(cell.value).strip().lower() == "true":
                cell.fill = GREEN_FILL
            elif cell.value is not None and str(cell.value).strip() != "":
                cell.fill = RED_FILL

    for col_idx, column in enumerate(df.columns, start=1):
        values = [str(column)] + [str(v) for v in df[column].tolist() if v is not None]
        width = min(max(len(v) for v in values) + 2, MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = width


def _add_energy_chart(worksheet, df: pd.DataFrame):
    """Bar chart of energy per case, anchored to the right of the table."""
    if "energy_eV" not in df.columns or "case" not in df.columns:
        return
    energies = pd.to_numeric(df["energy_eV"], errors="coerce")
    if energies.notna().sum() < 2:
        return

    energy_col = df.columns.get_loc("energy_eV") + 1
    case_col = df.columns.get_loc("case") + 1
    nrows = len(df)

    chart = BarChart()
    chart.title = "Energy per case"
    chart.y_axis.title = "energy (eV)"
    chart.legend = None
    chart.add_data(
        Reference(worksheet, min_col=energy_col, min_row=1, max_row=nrows + 1),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(worksheet, min_col=case_col, min_row=2, max_row=nrows + 1)
    )
    anchor_col = worksheet.cell(row=1, column=len(df.columns) + 2).column_letter
    worksheet.add_chart(chart, f"{anchor_col}2")


def write_results_to_excel(excel_path: str, rows: list[dict]):
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        df = pd.DataFrame(columns=COLUMN_ORDER)
    else:
        df = pd.DataFrame(rows)

    ordered_cols = [c for c in COLUMN_ORDER if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + remaining_cols]

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="results")
        _style_worksheet(writer.sheets["results"], df)
        _add_energy_chart(writer.sheets["results"], df)

    print(f"[EXCEL] Summary written to {excel_path}")
